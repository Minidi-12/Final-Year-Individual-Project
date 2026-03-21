# Data Loading Script

import pandas as pd
import os
from sqlalchemy import create_engine, text
from openpyxl import load_workbook
from dotenv import load_dotenv
from urllib.parse import quote_plus

load_dotenv()

# ── DATABASE CONNECTION ──────────────────────────────────────────
print("Connecting to mhealth_lk database...")

password = quote_plus(os.getenv('DB_PASSWORD'))
engine = create_engine(
    f"postgresql://{os.getenv('DB_USER')}:"
    f"{password}@"
    f"{os.getenv('DB_HOST')}:"
    f"{os.getenv('DB_PORT')}/"
    f"{os.getenv('DB_NAME')}"
)

try:
    with engine.connect() as conn:
        conn.execute(text('SELECT 1'))
    print(" Connected to database successfully!")
except Exception as e:
    print(f" Connection failed: {e}")
    exit()

# STEP 1: Load 25 Districts from Your Excel File

print("\n--- STEP 1: Loading Districts ---")

PROVINCE_MAP = {
    'Colombo': 'Western',        'Gampaha': 'Western',
    'Kalutara': 'Western',       'Kandy': 'Central',
    'Matale': 'Central',         'Nuwara Eliya': 'Central',
    'Galle': 'Southern',         'Matara': 'Southern',
    'Hambantota': 'Southern',    'Jaffna': 'Northern',
    'Kilinochchi': 'Northern',   'Mannar': 'Northern',
    'Vavuniya': 'Northern',      'Mullaitivu': 'Northern',
    'Batticaloa': 'Eastern',     'Ampara': 'Eastern',
    'Trincomalee': 'Eastern',    'Kurunegala': 'North Western',
    'Puttalam': 'North Western', 'Anuradhapura': 'North Central',
    'Polonnaruwa': 'North Central', 'Badulla': 'Uva',
    'Monaragala': 'Uva',         'Ratnapura': 'Sabaragamuwa',
    'Kegalle': 'Sabaragamuwa'
}

try:
    # ── Clear existing data first to avoid duplicates ──────────
    with engine.connect() as conn:
        conn.execute(text(
            'TRUNCATE TABLE vulnerability_scores CASCADE'
        ))
        conn.execute(text(
            'TRUNCATE TABLE districts CASCADE'
        ))
        conn.commit()
    print("   Cleared existing districts and scores")

    # ── Read Excel ─────────────────────────────────────────────
    wb = load_workbook(
        'data/District_Data_Extraction.xlsx',
        read_only=True
    )
    ws = wb['Sheet2']
    rows = list(ws.iter_rows(values_only=True))

    districts_data = []
    for row in rows[1:]:
        if row[0] is None:
            continue
        name = row[0]

        female_vals = [
            v for v in row[2:10]
            if isinstance(v, (int, float))
        ]
        f_total = int(sum(female_vals))

        districts_data.append({
            'name':                name,
            'province':            PROVINCE_MAP.get(name, 'Unknown'),
            'total_population':    int(row[1]) if row[1] else None,
            'female_10_14':        int(row[2]) if row[2] else None,
            'female_15_19':        int(row[3]) if row[3] else None,
            'female_20_24':        int(row[4]) if row[4] else None,
            'female_25_29':        int(row[5]) if row[5] else None,
            'female_30_34':        int(row[6]) if row[6] else None,
            'female_35_39':        int(row[7]) if row[7] else None,
            'female_40_44':        int(row[8]) if row[8] else None,
            'female_45_49':        int(row[9]) if row[9] else None,
            'female_10_49_total':  f_total,
            'mean_income_lkr':     float(row[11]) if row[11] else None,
            'poverty_rate_pct':    float(row[12]) if row[12] else None,
            'female_literacy_pct': float(row[13]) if row[13] else None,
        })

    df_districts = pd.DataFrame(districts_data)
    df_districts.to_sql(
        'districts', engine,
        if_exists='append', index=False
    )

    count = pd.read_sql(
        'SELECT COUNT(*) FROM districts', engine
    ).iloc[0, 0]
    print(f" Loaded {count} districts (expected 25)")
    print(df_districts[[
        'name', 'province',
        'total_population',
        'mean_income_lkr',
        'poverty_rate_pct'
    ]].to_string())

except Exception as e:
    print(f" District loading failed: {e}")
    import traceback
    traceback.print_exc()

# STEP 2: Create Vulnerability Score Placeholders

print("\n--- STEP 2: Creating Vulnerability Score Placeholders ---")

try:
    dist_ids = pd.read_sql(
        'SELECT id, name FROM districts', engine
    )
    with engine.connect() as conn:
        for _, row in dist_ids.iterrows():
            conn.execute(text('''
                INSERT INTO vulnerability_scores (district_id)
                VALUES (:did)
                ON CONFLICT (district_id) DO NOTHING
            '''), {'did': int(row['id'])})
        conn.commit()

    count = pd.read_sql(
        'SELECT COUNT(*) FROM vulnerability_scores', engine
    ).iloc[0, 0]
    print(f" Created {count} placeholder rows (expected 25)")

except Exception as e:
    print(f" Vulnerability scores setup failed: {e}")


# STEP 3: Load Survey Responses

print("\n--- STEP 3: Loading Survey Responses ---")

try:
    # ── Clear existing survey data ─────────────────────────────
    with engine.connect() as conn:
        conn.execute(text('TRUNCATE TABLE survey_responses CASCADE'))
        conn.commit()
    print("   Cleared existing survey responses")

    surveys = pd.read_excel(
        'data/Menstrual Health Access Survey (Responses).xlsx'
    )
    print(f"   Found {len(surveys)} rows in survey file")
    print(f"   Columns found: {list(surveys.columns[:5])}...")

    # ── Map district names to IDs ──────────────────────────────
    dist_map = dict(
        pd.read_sql('SELECT name, id FROM districts', engine).values
    )

    # ── FIX 1: Clean district names with Sinhala text ─────────
    def clean_district(name):
        if pd.isna(name):
            return None
        # Removes Sinhala text in brackets
        # e.g. 'Colombo (කොළඹ)' → 'Colombo'
        name = str(name).strip()
        if '(' in name:
            name = name[:name.index('(')].strip()
        return name

    district_col = [
        c for c in surveys.columns
        if 'Q2' in c or 'district' in c.lower()
    ]

    if district_col:
        surveys['district_id'] = (
            surveys[district_col[0]]
            .apply(clean_district)
            .map(dist_map)
        )
        unmapped = surveys[surveys['district_id'].isna()][
            district_col[0]
        ].unique()
        if len(unmapped) > 0:
            print(f"    Unmapped districts: {unmapped}")
        else:
            print("    All districts mapped successfully")
    else:
        surveys['district_id'] = None
        print("    District column not found")

    # ── FIX 2: Map columns — handle duplicates properly ───────
    col_map = {}
    used_names = {}

    for col in surveys.columns:
        mapped = None

        if 'Timestamp' in col:
            mapped = 'submitted_at'
        elif 'Q1:' in col and 'age' in col.lower():
            mapped = 'age_group'
        elif 'Q3:' in col:
            mapped = 'area_type'
        elif 'Q4:' in col:
            mapped = 'monthly_household_income'
        elif 'Q5:' in col:
            mapped = 'education_level'
        elif 'Q6:' in col:
            mapped = 'menstruating_count'
        elif 'Q7:' in col:
            mapped = 'product_type'
        elif 'Q8:' in col:
            mapped = 'monthly_spend_lkr'
        elif 'Q9:' in col:
            mapped = 'purchase_location'
        elif 'Q10:' in col:
            mapped = 'shop_name'
        elif 'Q11:' in col:
            mapped = 'shop_distance'
        elif 'Q12:' in col:
            mapped = 'affordability_issues_6m'
        elif 'Q13:' in col:
            mapped = 'coping_strategy'
        elif 'Q14:' in col:
            mapped = 'products_too_expensive'
        elif 'Q15:' in col:
            mapped = 'availability_rating'
        elif 'Q16:' in col:
            mapped = 'comfortable_discussing'
        # ── FIX 3: Q17 has 4 sub-questions — map each separately
        elif 'Q17:' in col and 'religious places' in col.lower():
            mapped = 'avoids_religious_places'
        elif 'Q17:' in col and 'ceremonies' in col.lower():
            mapped = 'avoids_religious_ceremonies'
        elif 'Q17:' in col and 'social' in col.lower():
            mapped = 'avoids_social_events'
        elif 'Q17:' in col and 'exercise' in col.lower():
            mapped = 'avoids_physical_exercise'
        elif 'Q18' in col:
            mapped = 'experienced_discrimination'
        elif 'Q19:' in col:
            mapped = 'community_impurity_belief'
        elif 'Q20:' in col:
            mapped = 'embarrassed_buying'
        elif 'Q21:' in col:
            mapped = 'age_first_learned'
        elif 'Q22:' in col:
            mapped = 'who_taught'
        elif 'Q23:' in col:
            mapped = 'received_school_education'
        elif 'Q24:' in col:
            mapped = 'pad_change_knowledge'
        elif 'Q25:' in col:
            mapped = 'knows_hygiene_risks'
        elif 'Q26:' in col:
            mapped = 'misses_school_work'
        elif 'Q27:' in col:
            mapped = 'miss_reason'
        elif 'Q30:' in col:
            mapped = 'daily_life_impact'

        if mapped:
            # Only map first occurrence — skip duplicates
            if mapped not in used_names:
                col_map[col] = mapped
                used_names[mapped] = col
            else:
                print(f"    Skipping duplicate: {col} → {mapped}")

    surveys_clean = surveys.rename(columns=col_map)

    # ── FIX 4: valid_cols includes all Q17 sub-columns ────────
    valid_cols = [
        'submitted_at', 'district_id', 'age_group',
        'area_type', 'monthly_household_income',
        'education_level', 'menstruating_count',
        'product_type', 'monthly_spend_lkr',
        'purchase_location', 'shop_name', 'shop_distance',
        'affordability_issues_6m', 'coping_strategy',
        'products_too_expensive', 'availability_rating',
        'comfortable_discussing',
        'avoids_religious_places',
        'avoids_religious_ceremonies',
        'avoids_social_events',
        'avoids_physical_exercise',
        'experienced_discrimination',
        'community_impurity_belief', 'embarrassed_buying',
        'age_first_learned', 'who_taught',
        'received_school_education', 'pad_change_knowledge',
        'knows_hygiene_risks', 'misses_school_work',
        'miss_reason', 'daily_life_impact'
    ]

    existing_cols = [
        c for c in valid_cols
        if c in surveys_clean.columns
    ]
    surveys_clean = surveys_clean[existing_cols]

    # ── Fix timestamp ──────────────────────────────────────────
    if 'submitted_at' in surveys_clean.columns:
        surveys_clean['submitted_at'] = pd.to_datetime(
            surveys_clean['submitted_at'], errors='coerce'
        )

    # ── Fix availability_rating to integer ────────────────────
    if 'availability_rating' in surveys_clean.columns:
        surveys_clean['availability_rating'] = pd.to_numeric(
            surveys_clean['availability_rating'], errors='coerce'
        ).astype('Int64')

    # ── Fix menstruating_count to integer ─────────────────────
    if 'menstruating_count' in surveys_clean.columns:
        surveys_clean['menstruating_count'] = pd.to_numeric(
            surveys_clean['menstruating_count'], errors='coerce'
        ).astype('Int64')

    surveys_clean.to_sql(
        'survey_responses', engine,
        if_exists='append', index=False
    )

    count = pd.read_sql(
        'SELECT COUNT(*) FROM survey_responses', engine
    ).iloc[0, 0]
    print(f" Loaded {count} survey responses")

    # ── Show district breakdown ────────────────────────────────
    district_breakdown = pd.read_sql('''
        SELECT d.name, COUNT(s.id) as responses
        FROM districts d
        LEFT JOIN survey_responses s
        ON s.district_id = d.id
        GROUP BY d.name
        ORDER BY responses DESC
    ''', engine)
    print("\n   Survey responses by district:")
    print(district_breakdown.to_string(index=False))

except Exception as e:
    print(f" Survey loading failed: {e}")
    import traceback
    traceback.print_exc()

# STEP 4: Load YouTube Comments

print("\n--- STEP 4: Loading YouTube Comments ---")

try:
    # ── Clear existing comments ────────────────────────────────
    with engine.connect() as conn:
        conn.execute(text('TRUNCATE TABLE youtube_comments CASCADE'))
        conn.commit()
    print("   Cleared existing YouTube comments")

    comments = pd.read_excel(
        'data/YouTube_Comments_Cleaned.xlsx'
    )
    print(f"   Found {len(comments)} comments")
    print(f"   Columns: {list(comments.columns)}")

    # ── Build clean dataframe ──────────────────────────────────
    comments_clean = pd.DataFrame({
        'author':
            comments['author'] if 'author' in comments.columns
            else '',
        'comment_text':
            comments['comment'] if 'comment' in comments.columns
            else '',
        'page_url':
            comments['pageUrl'] if 'pageUrl' in comments.columns
            else '',
        'video_title':
            comments['title'] if 'title' in comments.columns
            else '',
        'sentiment_label':
            comments['Sentimnet'] if 'Sentimnet' in comments.columns
            else '',
        'topic_raw':
            comments['Topic'] if 'Topic' in comments.columns
            else '',
    })

    # ── Clean sentiment labels ─────────────────────────────────
    comments_clean['sentiment_label'] = (
        comments_clean['sentiment_label']
        .astype(str)
        .str.strip()
        .replace({'nan': None, '': None, ' ': None})
    )

    # ── Clean topic labels ─────────────────────────────────────
    comments_clean['topic_raw'] = (
        comments_clean['topic_raw']
        .astype(str)
        .str.strip()
        .replace({'nan': None, '': None, ' ': None})
    )

    comments_clean.to_sql(
        'youtube_comments', engine,
        if_exists='append', index=False
    )

    count = pd.read_sql(
        'SELECT COUNT(*) FROM youtube_comments', engine
    ).iloc[0, 0]
    print(f" Loaded {count} YouTube comments")

    # ── Show sentiment distribution ────────────────────────────
    sentiment_dist = pd.read_sql('''
        SELECT sentiment_label, COUNT(*) as count
        FROM youtube_comments
        WHERE sentiment_label IS NOT NULL
        GROUP BY sentiment_label
        ORDER BY count DESC
    ''', engine)
    print("\n   Sentiment Distribution:")
    print(sentiment_dist.to_string(index=False))

except Exception as e:
    print(f" YouTube comments loading failed: {e}")
    import traceback
    traceback.print_exc()

# FINAL VERIFICATION

print("\n" + "=" * 50)
print("DATABASE LOADING COMPLETE — FINAL CHECK")
print("=" * 50)

tables = [
    ('districts',            25),
    ('vulnerability_scores', 25),
    ('survey_responses',    347),
    ('youtube_comments',   1247),
    ('retail_outlets',        0),
    ('users',                 0),
    ('cycle_logs',            0),
]

all_good = True
for table, expected in tables:
    try:
        count = pd.read_sql(
            f'SELECT COUNT(*) FROM {table}', engine
        ).iloc[0, 0]
        status = (
            "✅" if (expected == 0 or count >= expected * 0.95)
            else "⚠️"
        )
        if status == "⚠️":
            all_good = False
        print(
            f"  {status} {table}: {count} rows "
            f"(expected ~{expected})"
        )
    except Exception as e:
        print(f"   {table}: could not check — {e}")
        all_good = False

print("\n" + (
    " ALL DATA LOADED SUCCESSFULLY!"
    if all_good
    else " Some tables need attention — check above"
))